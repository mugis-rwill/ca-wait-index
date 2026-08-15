"""
ingest.py — load and clean the raw CIHI "Wait Times for Priority Procedures" table.

Expected input: a CSV export of CIHI's Table 1 (Wait times for priority
procedures, by province and Canada) with these columns, matching CIHI's
published long/tidy format:

    Reporting level, Province, Region, Indicator, Metric,
    Data year, Unit of measurement, Indicator result

- Reporting level: "Provincial" | "National" | "Regional"
- Province: province name, or "Canada" for national rows
- Region: health region/zone name, or blank/"n/a" for provincial rows
- Indicator: procedure name (e.g. "Hip Replacement", "Knee Replacement")
- Metric: "50th percentile" | "90th percentile" | "Volume" | "% meeting benchmark"
- Data year: "2020", "2020FY", "2020Q3Q4", etc.
- Indicator result: numeric value, or blank/"N/A"

This module does NOT compute anything diagnostic (no regression, no
residuals) — it only ingests and reshapes. All statistics live in compute.py,
so the two stay independently testable and the math is never duplicated
in the frontend JS.
"""

import re
import pandas as pd
import openpyxl

PROCEDURES = ["Hip Replacement", "Knee Replacement"]

METRIC_MAP = {
    "50th percentile": "wait50",
    "90th percentile": "wait90",
    "Volume": "volume",
    "% meeting benchmark": "pct_benchmark",
}

# Provinces we report on. Territories are excluded, consistent with CIHI's
# own wait-time reporting, which excludes them for data-consistency reasons.
PROVINCES = [
    "British Columbia", "Alberta", "Saskatchewan", "Manitoba", "Ontario",
    "Quebec", "New Brunswick", "Nova Scotia", "Prince Edward Island",
    "Newfoundland and Labrador",
]


def extract_year(raw_year) -> int | None:
    """Normalize 'Data year' values like '2020', '2020FY', '2020Q3Q4' -> 2020."""
    if pd.isna(raw_year):
        return None
    match = re.match(r"^(\d{4})", str(raw_year).strip())
    return int(match.group(1)) if match else None


def load_raw(xlsx_path: str, sheet_name: str | None = None) -> pd.DataFrame:
    """
    Load CIHI's 'Wait times for priority procedures, by province and Canada'
    sheet. Different CIHI editions have used different sheet names ('Table 1',
    'Wait times 2008 to 2021', etc.) and slightly different header offsets,
    so instead of hardcoding row positions, this scans for the header row
    (the one starting with 'Reporting level') and reads everything after it.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_name = sheet_name or next(
        n for n in wb.sheetnames if n.lower().startswith("table") or "wait times" in n.lower()
    )
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)

    header = None
    for r in rows:
        if r and isinstance(r[0], str) and r[0].strip() == "Reporting level":
            header = r
            break
    if header is None:
        raise ValueError(f"Could not find 'Reporting level' header row in sheet '{sheet_name}'")

    data = [r[:8] for r in rows if r and r[0] is not None]
    df = pd.DataFrame(data, columns=[
        "Reporting level", "Province", "Region", "Indicator",
        "Metric", "Data year", "Unit of measurement", "Indicator result",
    ])
    return df


def clean(df: pd.DataFrame, level: str = "Provincial") -> pd.DataFrame:
    """
    Filter to hip/knee rows at the given reporting level and coerce types.
    "Provincial" rows have no sub-region; "Regional" rows have one — used to
    pull CIHI's health-region breakdown for the drill-down view, since
    StatCan has no sub-provincial population data to normalize it against.
    """
    df = df.copy()
    df = df[df["Reporting level"].isin(["Provincial", "National", "Regional"])]
    df["year"] = df["Data year"].apply(extract_year)

    is_level = df["Reporting level"].astype(str).str.strip().eq(level)
    is_target_procedure = df["Indicator"].astype(str).str.strip().isin(PROCEDURES)
    is_named_province = df["Province"].astype(str).str.strip().isin(PROVINCES)
    # Provincial rows should have no sub-region; regional rows should. Treat
    # blank/NaN/"n/a" as "no region".
    region = df["Region"].fillna("n/a").astype(str).str.strip().str.lower()
    has_no_region = region.isin(["", "n/a", "na"])
    region_matches = ~has_no_region if level == "Regional" else has_no_region

    df = df[is_level & is_target_procedure & is_named_province & region_matches]

    df["metric_col"] = df["Metric"].astype(str).str.strip().map(METRIC_MAP)
    df = df[df["metric_col"].notna()]

    df["Indicator result"] = pd.to_numeric(
        df["Indicator result"].replace({"N/A": None, "n/a": None, "": None}),
        errors="coerce",
    )
    return df


def to_wide(df: pd.DataFrame, year: int, by_region: bool = False) -> pd.DataFrame:
    """
    Pivot to one row per (province[, region], procedure) for a given year,
    with wait50 / wait90 / volume / pct_benchmark as columns.
    """
    year_df = df[df["year"] == year]
    if year_df.empty:
        return pd.DataFrame()

    index_cols = ["Province", "Region", "Indicator"] if by_region else ["Province", "Indicator"]
    wide = year_df.pivot_table(
        index=index_cols,
        columns="metric_col",
        values="Indicator result",
        aggfunc="first",
    ).reset_index()
    rename = {"Province": "province", "Indicator": "procedure"}
    if by_region:
        rename["Region"] = "region"
    wide = wide.rename(columns=rename)
    for col in ["wait50", "wait90", "volume", "pct_benchmark"]:
        if col not in wide.columns:
            wide[col] = None
    return wide


POPULATION_AGE_GROUP = "65 years and older"


def load_population(csv_path: str) -> pd.DataFrame:
    """
    Load StatCan Table 17-10-0005-01 (population estimates by age/gender),
    filtered to the 65+ total for each province/year. One row per
    (province, year).
    """
    df = pd.read_csv(csv_path)
    df = df[
        (df["Gender"].astype(str).str.strip() == "Total - gender")
        & (df["Age group"].astype(str).str.strip() == POPULATION_AGE_GROUP)
        & (df["GEO"].astype(str).str.strip().isin(PROVINCES))
    ]
    out = df[["GEO", "REF_DATE", "VALUE"]].rename(
        columns={"GEO": "province", "REF_DATE": "year", "VALUE": "population"}
    )
    out["year"] = out["year"].astype(int)
    out["population"] = out["population"].astype(float)
    return out.reset_index(drop=True)


def population_growth(pop_df: pd.DataFrame) -> pd.DataFrame:
    """
    Year-over-year growth rate (in percentage points) of the 65+ population
    per province — used as a real demand-pressure proxy. The first year in
    the table has no prior-year baseline, so it's dropped rather than given
    a fabricated growth rate.
    """
    pop_df = pop_df.sort_values(["province", "year"]).copy()
    pop_df["prev_population"] = pop_df.groupby("province")["population"].shift(1)
    pop_df["growth_pct"] = (
        (pop_df["population"] - pop_df["prev_population"]) / pop_df["prev_population"] * 100
    )
    return pop_df.dropna(subset=["growth_pct"])[["province", "year", "population", "growth_pct"]]
